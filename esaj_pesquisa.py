#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import time
import sys
import os
import re
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl nao instalado. Execute setup.bat primeiro.")
    input("Pressione Enter para sair...")
    sys.exit(1)

try:
    from selenium import webdriver
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, WebDriverException
    )
except ImportError:
    print("ERRO: selenium nao instalado. Execute setup.bat primeiro.")
    input("Pressione Enter para sair...")
    sys.exit(1)


ARQUIVO_ENTRADA = "CONSULTA PROCESSOS 2026 - COPIA.xlsx"
ARQUIVO_SAIDA = "RESULTADOS_CONSULTA.xlsx"
DEBUGGER_PORT = 9222

COLUNAS = [
    "Nº PROCESSO",
    "CLASSE",
    "ASSUNTO",
    "FORO",
    "VARA",
    "JUIZ/JUÍZA",
    "VALOR DA AÇÃO",
    "SITUAÇÃO",
    "PARTE AUTORA/REQUERENTE",
    "PARTE RÉ/REQUERIDA",
    "ÚLTIMO ANDAMENTO",
    "DATA ÚLTIMO ANDAMENTO",
    "STATUS DA PESQUISA",
]

URL_1GRAU = "https://esaj.tjsp.jus.br/cpopg/open.do"
URL_2GRAU = "https://esaj.tjsp.jus.br/cposg/open.do"


def conectar_edge():
    options = EdgeOptions()
    options.add_experimental_option("debuggerAddress", f"localhost:{DEBUGGER_PORT}")
    try:
        driver = webdriver.Edge(options=options)
        print("  Conectado ao Edge com sucesso!")
        return driver
    except WebDriverException as e:
        print(f"  ERRO ao conectar ao Edge: {e}")
        print("  Certifique-se de ter executado START_EDGE-MANUAL_LOGIN.bat antes!")
        return None


def ler_processos():
    if not os.path.exists(ARQUIVO_ENTRADA):
        print(f"ERRO: Arquivo '{ARQUIVO_ENTRADA}' nao encontrado!")
        return []

    wb = openpyxl.load_workbook(ARQUIVO_ENTRADA)
    ws = wb.active

    processos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        valor = row[0] if row else None
        if valor and str(valor).strip():
            texto = str(valor).strip()
            # Ignorar linhas de instrução ou exemplo
            if texto.lower().startswith(("instrução", "instrucao", "exemplo")):
                continue
            processos.append(texto)

    print(f"  {len(processos)} processo(s) encontrado(s) para pesquisar.")
    return processos


def normalizar_numero(numero):
    """Tenta formatar número para padrão CNJ NNNNNNN-DD.AAAA.J.TT.OOOO."""
    digitos = re.sub(r"[^0-9]", "", numero)
    if len(digitos) == 20:
        d = digitos
        return f"{d[0:7]}-{d[7:9]}.{d[9:13]}.{d[13]}.{d[14:16]}.{d[16:20]}"
    return numero


def detectar_grau(numero):
    """Detecta se o processo é de 1º ou 2º grau pelo número CNJ."""
    # 2º grau no TJSP: segmento TT = 26 e origem iniciando em 0000 ou tipo de classe
    # Heurística: processos com origem 0000 geralmente são 2º grau
    partes = re.split(r"[-.]", numero)
    if len(partes) >= 5:
        origem = partes[-1]
        if origem == "0000":
            return 2
    return 1


def _texto_elemento(driver, seletores):
    """Tenta múltiplos seletores e retorna o texto do primeiro que encontrar."""
    for by, valor in seletores:
        try:
            el = driver.find_element(by, valor)
            texto = el.text.strip()
            if texto:
                return texto
        except NoSuchElementException:
            pass
    return ""


def extrair_partes(driver):
    """Extrai partes do processo (autores e réus)."""
    autores, reus = [], []
    try:
        tabela = driver.find_element(By.ID, "tableTodasPartes")
        linhas = tabela.find_elements(By.TAG_NAME, "tr")
        role_atual = ""
        for linha in linhas:
            tds = linha.find_elements(By.TAG_NAME, "td")
            if len(tds) >= 2:
                role_td = tds[0].text.strip().lower()
                nome_td = tds[1].text.strip()
                if role_td:
                    if any(p in role_td for p in ["autor", "requerente", "exequente", "reclamante", "impetrante"]):
                        role_atual = "autor"
                    elif any(p in role_td for p in ["réu", "reu", "requerido", "executado", "reclamado", "impetrado"]):
                        role_atual = "reu"
                    elif any(p in role_td for p in ["advogad", "curador", "inventariante"]):
                        role_atual = ""
                if nome_td and role_atual == "autor":
                    autores.append(nome_td.split("\n")[0])
                elif nome_td and role_atual == "reu":
                    reus.append(nome_td.split("\n")[0])
    except NoSuchElementException:
        pass

    return (
        " | ".join(autores[:3]) if autores else "",
        " | ".join(reus[:3]) if reus else "",
    )


def extrair_ultimo_andamento(driver):
    """Retorna (data, descricao) do último andamento."""
    seletores_tabela = [
        "tabelaTodasMovimentacoes",
        "tabelaUltimasMovimentacoes",
    ]
    for id_tabela in seletores_tabela:
        try:
            tabela = driver.find_element(By.ID, id_tabela)
            linhas = tabela.find_elements(By.XPATH, ".//tr[not(contains(@class,'fundoAzulEscuro'))]")
            for linha in linhas:
                tds = linha.find_elements(By.TAG_NAME, "td")
                if len(tds) >= 2:
                    data = tds[0].text.strip()
                    descricao = tds[1].text.strip()
                    if data and descricao:
                        return data, descricao[:300]
        except NoSuchElementException:
            pass
    return "", ""


def inferir_situacao(ultimo_andamento):
    """Infere situação do processo com base no último andamento."""
    texto = ultimo_andamento.lower()
    if any(p in texto for p in ["arquivad", "arquivo definitivo"]):
        return "Arquivado"
    if any(p in texto for p in ["extinto", "extinção"]):
        return "Extinto"
    if any(p in texto for p in ["sentença", "sentenca", "julgado procedente", "julgado improcedente"]):
        return "Sentenciado"
    if any(p in texto for p in ["acórdão", "acordao"]):
        return "Julgado (2º Grau)"
    if any(p in texto for p in ["baixad", "transitou em julgado"]):
        return "Baixado"
    return "Em andamento"


def pesquisar_processo(driver, numero_processo):
    resultado = {col: "" for col in COLUNAS}
    resultado["Nº PROCESSO"] = numero_processo

    numero_norm = normalizar_numero(numero_processo)
    grau = detectar_grau(numero_norm)
    url_base = URL_1GRAU if grau == 1 else URL_2GRAU

    try:
        driver.get(url_base)
        time.sleep(2)

        wait = WebDriverWait(driver, 20)

        # Preencher número unificado CNJ
        try:
            campo_num = wait.until(
                EC.presence_of_element_located((By.ID, "numeroDigitoAnoUnificado"))
            )
            campo_num.clear()
            # Parte NNNNNNN-DD
            match = re.match(r"(\d{7}-\d{2})\.(\d{4})\.(\d)\.(\d{2})\.(\d{4})", numero_norm)
            if match:
                campo_num.send_keys(match.group(1))
                # Ano
                try:
                    driver.find_element(By.ID, "anoDig").clear()
                    driver.find_element(By.ID, "anoDig").send_keys(match.group(2))
                except NoSuchElementException:
                    pass
                # Foro (último grupo)
                try:
                    campo_foro = driver.find_element(By.ID, "foroNumeroUnificado")
                    campo_foro.clear()
                    campo_foro.send_keys(match.group(5))
                except NoSuchElementException:
                    pass
            else:
                campo_num.send_keys(numero_norm)
        except TimeoutException:
            # Fallback: campo legado
            try:
                campo = driver.find_element(By.NAME, "processo.numero")
                campo.clear()
                campo.send_keys(numero_norm)
            except NoSuchElementException:
                resultado["STATUS DA PESQUISA"] = "Erro: campo não encontrado"
                return resultado

        # Clicar em Consultar
        try:
            driver.find_element(By.ID, "botaoConsultarProcessos").click()
        except NoSuchElementException:
            try:
                driver.find_element(
                    By.XPATH, "//input[@value='Consultar' or @value='Pesquisar']"
                ).click()
            except NoSuchElementException:
                driver.find_element(
                    By.XPATH, "//button[contains(., 'Consultar')]"
                ).click()

        time.sleep(3)

        page = driver.page_source
        if any(
            msg in page
            for msg in [
                "Não existe processo",
                "não foi encontrado",
                "Processo não encontrado",
                "nenhum processo",
            ]
        ):
            resultado["STATUS DA PESQUISA"] = "Não encontrado"
            return resultado

        # --- Extração de dados ---
        resultado["CLASSE"] = _texto_elemento(
            driver,
            [
                (By.ID, "classeProcesso"),
                (By.XPATH, "//span[@id='classeProcesso']"),
                (By.XPATH, "//td[contains(@id,'Classe')]"),
            ],
        )
        resultado["ASSUNTO"] = _texto_elemento(
            driver,
            [
                (By.ID, "assuntoProcesso"),
                (By.XPATH, "//td[contains(@id,'Assunto')]"),
                (By.XPATH, "//span[contains(text(),'Assunto')]/following-sibling::span[1]"),
            ],
        )
        resultado["FORO"] = _texto_elemento(
            driver,
            [
                (By.ID, "foroProcesso"),
                (By.XPATH, "//span[contains(text(),'Foro')]/following-sibling::span[1]"),
            ],
        )
        resultado["VARA"] = _texto_elemento(
            driver,
            [
                (By.ID, "varaProcesso"),
                (By.XPATH, "//span[contains(text(),'Vara')]/following-sibling::span[1]"),
            ],
        )
        resultado["JUIZ/JUÍZA"] = _texto_elemento(
            driver,
            [
                (By.ID, "juizProcesso"),
                (By.XPATH, "//span[contains(text(),'Juiz')]/following-sibling::span[1]"),
                (By.XPATH, "//span[contains(text(),'Magistrad')]/following-sibling::span[1]"),
            ],
        )
        resultado["VALOR DA AÇÃO"] = _texto_elemento(
            driver,
            [
                (By.ID, "valorAcaoProcesso"),
                (By.XPATH, "//span[contains(text(),'Valor')]/following-sibling::span[1]"),
            ],
        )

        autor, reu = extrair_partes(driver)
        resultado["PARTE AUTORA/REQUERENTE"] = autor
        resultado["PARTE RÉ/REQUERIDA"] = reu

        data_and, desc_and = extrair_ultimo_andamento(driver)
        resultado["DATA ÚLTIMO ANDAMENTO"] = data_and
        resultado["ÚLTIMO ANDAMENTO"] = desc_and

        situacao = _texto_elemento(
            driver,
            [
                (By.XPATH, "//span[contains(@id,'situacao') or contains(@id,'Situacao')]"),
                (By.XPATH, "//td[contains(@id,'situacao')]"),
            ],
        )
        resultado["SITUAÇÃO"] = situacao if situacao else inferir_situacao(desc_and)

        resultado["STATUS DA PESQUISA"] = "Encontrado"

    except Exception as exc:
        resultado["STATUS DA PESQUISA"] = f"Erro: {str(exc)[:80]}"

    return resultado


def salvar_resultados(resultados):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    azul_escuro = "0B2740"
    azul_medio = "1D5F96"
    verde_claro = "E8F5E9"
    amarelo_claro = "FFF9C4"
    vermelho_claro = "FFEBEE"
    cinza_claro = "F5F8FC"

    borda_fina = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Linha de título (linha 1)
    n_cols = len(COLUNAS)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    cel = ws["A1"]
    cel.value = f"RESULTADOS DA CONSULTA ESAJ  —  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cel.font = Font(bold=True, size=14, color="FFFFFF")
    cel.fill = PatternFill(start_color=azul_escuro, end_color=azul_escuro, fill_type="solid")
    cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Cabeçalhos (linha 2)
    for idx, nome in enumerate(COLUNAS, 1):
        cel = ws.cell(row=2, column=idx, value=nome)
        cel.font = Font(bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill(start_color=azul_medio, end_color=azul_medio, fill_type="solid")
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = borda_fina
    ws.row_dimensions[2].height = 36

    # Dados
    for r_idx, dados in enumerate(resultados, 3):
        status = dados.get("STATUS DA PESQUISA", "")
        if status == "Encontrado":
            cor_linha = verde_claro
        elif status == "Não encontrado":
            cor_linha = amarelo_claro
        else:
            cor_linha = vermelho_claro

        for c_idx, col in enumerate(COLUNAS, 1):
            cel = ws.cell(row=r_idx, column=c_idx, value=dados.get(col, ""))
            cel.alignment = Alignment(vertical="center", wrap_text=True)
            cel.border = borda_fina
            if col == "STATUS DA PESQUISA":
                cel.fill = PatternFill(start_color=cor_linha, end_color=cor_linha, fill_type="solid")
                cel.font = Font(bold=True)
            else:
                bg = cinza_claro if r_idx % 2 == 0 else "FFFFFF"
                cel.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ws.row_dimensions[r_idx].height = 22

    # Larguras das colunas
    larguras = [28, 22, 30, 22, 22, 28, 16, 20, 38, 38, 55, 22, 22]
    for i, larg in enumerate(larguras[:n_cols], 1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(n_cols)}{len(resultados) + 2}"

    # Aba Sumário
    ws2 = wb.create_sheet("Sumário")
    encontrados = sum(1 for r in resultados if r.get("STATUS DA PESQUISA") == "Encontrado")
    nao_enc = sum(1 for r in resultados if r.get("STATUS DA PESQUISA") == "Não encontrado")
    erros = len(resultados) - encontrados - nao_enc

    ws2["A1"] = "SUMÁRIO DA PESQUISA"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = f"Data/Hora:               {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws2["A3"] = f"Total pesquisado:        {len(resultados)}"
    ws2["A4"] = f"Encontrados:             {encontrados}"
    ws2["A5"] = f"Não encontrados:         {nao_enc}"
    ws2["A6"] = f"Erros:                   {erros}"
    ws2.column_dimensions["A"].width = 40

    wb.save(ARQUIVO_SAIDA)
    print(f"  Salvo: {ARQUIVO_SAIDA}")


def main():
    sep = "=" * 56
    print(sep)
    print("   SISTEMA DE PESQUISA AUTOMATICA  —  ESAJ TJSP")
    print(sep)
    print()

    print("[1/4] Lendo lista de processos...")
    processos = ler_processos()
    if not processos:
        print(f"\nNenhum processo encontrado em '{ARQUIVO_ENTRADA}'.")
        input("Pressione Enter para sair...")
        return

    print("\n[2/4] Conectando ao Edge (porta de depuracao)...")
    driver = conectar_edge()
    if not driver:
        input("\nPressione Enter para sair...")
        return

    print("\n[3/4] Verificando sessao ESAJ...")
    try:
        url_atual = driver.current_url
        if "login" in url_atual.lower() or "sajcas" in url_atual.lower():
            print("  AVISO: Sessao nao identificada no ESAJ.")
            print("  Faca o login no Edge e pressione Enter para continuar.")
            input("  -> Pressione Enter apos o login: ")
    except Exception:
        pass

    print(f"\n[4/4] Pesquisando {len(processos)} processo(s)...")
    print("-" * 56)

    resultados = []
    for i, proc in enumerate(processos, 1):
        print(f"  [{i:>3}/{len(processos)}] {proc}", end="  ")
        sys.stdout.flush()
        resultado = pesquisar_processo(driver, proc)
        resultados.append(resultado)
        status = resultado.get("STATUS DA PESQUISA", "")
        print(f"-> {status}")
        if i < len(processos):
            time.sleep(2)

    print("-" * 56)
    print(f"\nSalvando resultados...")
    salvar_resultados(resultados)

    print()
    print(sep)
    print(f"  CONCLUIDO!  {len(resultados)} processo(s) pesquisado(s).")
    print(f"  Abra '{ARQUIVO_SAIDA}' para ver os resultados.")
    print(sep)
    print()
    input("Pressione Enter para fechar...")


if __name__ == "__main__":
    main()

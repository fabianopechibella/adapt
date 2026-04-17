#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Cria as planilhas modelo para o sistema de pesquisa ESAJ."""

import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl nao instalado. Execute setup.bat primeiro.")
    input("Pressione Enter para sair...")
    sys.exit(1)


def criar_planilha_entrada():
    nome = "CONSULTA PROCESSOS 2026 - COPIA.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos"

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Título
    ws.merge_cells("A1:C1")
    cel = ws["A1"]
    cel.value = "CONSULTA DE PROCESSOS 2026 — Cole os processos abaixo"
    cel.font = Font(bold=True, size=13, color="FFFFFF")
    cel.fill = PatternFill(start_color="0B2740", end_color="0B2740", fill_type="solid")
    cel.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Cabeçalhos
    headers = ["NÚMERO DO PROCESSO", "OBSERVAÇÃO", "PRIORIDADE"]
    for col, h in enumerate(headers, 1):
        cel = ws.cell(row=2, column=col, value=h)
        cel.font = Font(bold=True, color="FFFFFF", size=11)
        cel.fill = PatternFill(start_color="1D5F96", end_color="1D5F96", fill_type="solid")
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = borda
    ws.row_dimensions[2].height = 28

    # Linha de instrução
    ws.merge_cells("A3:C3")
    cel = ws["A3"]
    cel.value = "INSTRUÇÃO: Cole os números de processo na coluna A (um por linha)"
    cel.font = Font(italic=True, color="888888", size=10)
    cel.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 20

    # Linha de exemplo (cinza, para o usuário apagar)
    exemplos = [
        "1001234-56.2023.8.26.0100",
        "0098765-43.2024.8.26.0050",
    ]
    for i, ex in enumerate(exemplos, 4):
        cel = ws.cell(row=i, column=1, value=ex)
        cel.font = Font(color="AAAAAA", italic=True)
        cel.border = borda
        ws.cell(row=i, column=2).border = borda
        ws.cell(row=i, column=3).border = borda
        ws.row_dimensions[i].height = 20

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = "A3"

    wb.save(nome)
    print(f"  Criado: {nome}")


def criar_planilha_saida():
    nome = "RESULTADOS_CONSULTA.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws["A1"] = "Arquivo gerado automaticamente pelo RUM_ESAJ — nao editar"
    ws["A1"].font = Font(italic=True, color="AAAAAA")
    ws.column_dimensions["A"].width = 50

    wb.save(nome)
    print(f"  Criado: {nome}")


def main():
    print()
    print("=" * 50)
    print("  CRIANDO PLANILHAS MODELO — SISTEMA ESAJ")
    print("=" * 50)
    print()
    criar_planilha_entrada()
    criar_planilha_saida()
    print()
    print("=" * 50)
    print("  Planilhas criadas com sucesso!")
    print()
    print("  Proximos passos:")
    print("  1. Abra 'CONSULTA PROCESSOS 2026 - COPIA.xlsx'")
    print("  2. Apague os exemplos e cole seus processos")
    print("  3. Execute START_EDGE-MANUAL_LOGIN.bat")
    print("  4. Faca login no ESAJ")
    print("  5. Execute RUM_ESAJ.bat")
    print("=" * 50)
    print()
    input("Pressione Enter para fechar...")


if __name__ == "__main__":
    main()
